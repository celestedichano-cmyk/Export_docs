"""
dossier_extractor.py
Extrae datos de un Dossier de producto (.xlsx) de The Not Company
para autocompletar el Reporte de Fórmula y el Informe de Aditivos.

Uso:
    with open("dossier.xlsx", "rb") as f:
        data = extract_dossier(f.read())
"""

from openpyxl import load_workbook
import io
import re
import unicodedata


def _fmt_pct(value):
    """
    Formatea una fracción decimal (ej 0.0303) como porcentaje legible.
    Usa suficientes decimales para conservar al menos 2 cifras significativas,
    evitando perder valores muy chicos (ej. vitaminas en trazas).
    """
    if value is None:
        return ""
    pct = value * 100
    if pct == 0:
        return "0"
    # Calcular decimales necesarios para mostrar al menos 2 cifras significativas
    import math
    magnitud = math.floor(math.log10(abs(pct)))
    decimales = max(6, -magnitud + 1)
    decimales = min(decimales, 15)
    s = f"{pct:.{decimales}f}".rstrip("0").rstrip(".")
    return s if s else "0"


def _fmt_ins(value):
    """Formatea el número INS: quita '.0' final si es un float entero, deja texto tal cual."""
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _normalize(text):
    """Normaliza texto para comparación: sin tildes, minúsculas, sin espacios extra."""
    if not text:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.split())


# Pares de nombres que refieren al mismo ingrediente/aditivo pero se
# escriben distinto entre la tabla principal (a veces nombre comercial,
# ej. "Betanina") y la tabla de orden decreciente (a veces nombre
# genérico/INS, ej. "Rojo de remolacha") — o viceversa. Se puede ampliar
# esta lista a medida que aparezcan más casos así en otros Dossiers.
GRUPOS_SINONIMOS = [
    {"rojo de remolacha", "betanina", "betanin"},
]
_SINONIMO_DE = {}
for _grupo in GRUPOS_SINONIMOS:
    for _nombre in _grupo:
        _SINONIMO_DE[_nombre] = _grupo


def _buscar_en_mapa(nombre, mapa):
    """
    Busca `nombre` en `mapa` (claves normalizadas), tolerando que la tabla
    principal y la de orden decreciente usen redacciones distintas para
    el mismo ingrediente:
    1) match exacto;
    2) un nombre contenido en el otro (ej. "metilcelulosa" está contenido
       en "metilcelulosa solucel") — con un largo mínimo para evitar falsos
       positivos con palabras cortas/genéricas;
    3) sinónimos conocidos sin relación textual (ej. "rojo de remolacha" /
       "betanina"), vía GRUPOS_SINONIMOS.
    Si más de una clave del mapa matchea por contención con valores
    distintos, se considera ambiguo y no se devuelve nada (mismo criterio
    de seguridad que el resto del matching de este proyecto).
    """
    n = _normalize(nombre)
    if not n:
        return ""

    if n in mapa:
        return mapa[n]

    if len(n) >= 5:
        candidatos = set()
        for clave, valor in mapa.items():
            if len(clave) >= 5 and (n in clave or clave in n):
                candidatos.add(valor)
        if len(candidatos) == 1:
            return next(iter(candidatos))

    grupo = _SINONIMO_DE.get(n)
    if grupo:
        candidatos = {mapa[alias] for alias in grupo if alias in mapa}
        if len(candidatos) == 1:
            return next(iter(candidatos))

    return ""


def _valor_como_porcentaje(cell):
    """
    Devuelve el valor de una celda numérica como porcentaje "plano" (ej.
    55.5759 significa 55.5759%), sin importar si el Dossier lo guarda como
    fracción con formato de celda "%" (valor crudo 0.555759) o como número
    ya escalado (valor crudo 55.5759) — distintos Dossiers usan una u otra
    convención según quién armó la plantilla.
    """
    if not isinstance(cell.value, (int, float)):
        return None
    if "%" in (cell.number_format or ""):
        return cell.value * 100
    return cell.value


def _resolve_formula_refs(formula, ws_values):
    """
    Dada una fórmula simple que solo suma referencias a celdas (ej '=R30'
    o '=V15+V16+V17'), devuelve la suma de los valores YA CALCULADOS
    (cacheados) de esas celdas, leídos desde ws_values (hoja cargada con
    data_only=True). Funciona sin importar a qué columna apunten las
    referencias — distintos Dossiers usan columnas de apoyo distintas
    (V, R, etc.) para esta cadena, según quién armó la plantilla.

    Devuelve None si la celda no es una fórmula o no contiene referencias
    de celda reconocibles.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    refs = re.findall(r'([A-Z]{1,3})(\d+)', formula.upper())
    if not refs:
        return None
    total = 0.0
    encontro_alguna = False
    for col_letters, row_str in refs:
        try:
            celda = ws_values[f"{col_letters}{row_str}"]
        except Exception:
            continue
        if isinstance(celda.value, (int, float)):
            total += celda.value
            encontro_alguna = True
    return total if encontro_alguna else None


def _buscar_encabezados(ws, etiquetas, max_filas=15, max_cols=40):
    """
    Busca, en las primeras `max_filas` filas de la hoja, una fila que
    contenga la mayoría de las `etiquetas` (texto normalizado, puede ser
    coincidencia parcial — "contiene") en distintas columnas de esa misma
    fila. Devuelve {etiqueta: columna} para esa fila, y el número de fila,
    o (None, None) si no se encuentra.

    Esto reemplaza la suposición anterior de que cada columna está en una
    posición fija (A, D, E, F, X, Y...) — los Dossiers de distintos
    copackers/equipos insertan columnas extra (ej. "Código SAP") que
    corren todo el resto de la tabla, así que hay que ubicar cada columna
    por el texto real de su encabezado.
    """
    mejor_fila = None
    mejor_mapa = {}
    for row in ws.iter_rows(min_row=1, max_row=max_filas, max_col=max_cols):
        mapa_fila = {}
        for cell in row:
            if not cell.value:
                continue
            texto = _normalize(cell.value)
            for etiqueta in etiquetas:
                if etiqueta not in mapa_fila and (texto == etiqueta or etiqueta in texto):
                    mapa_fila[etiqueta] = cell.column
        if len(mapa_fila) > len(mejor_mapa):
            mejor_mapa = mapa_fila
            mejor_fila = row[0].row
    return mejor_fila, mejor_mapa


def extract_dossier(xlsx_bytes):
    """
    Extrae producto, fórmula ordenada y datos de aditivos de un Dossier .xlsx.

    Returns:
        dict con:
        - 'producto'
        - 'formula_rows' (Ingrediente|% por línea, para Reporte Fórmula)
        - 'ingredientes' (uno por línea, orden decreciente, para Informe Aditivos)
        - 'aditivos' (NOMBRE | Función, uno por línea)
        - 'aditivos_filas' (números de fila 1-based que son aditivos, ej "8, 13")
    """
    result = {}
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    # --- Producto ---
    if "Datos de producto" in wb.sheetnames:
        ws_datos = wb["Datos de producto"]
        for row in ws_datos.iter_rows(min_row=1, max_row=30, max_col=2):
            label_cell, value_cell = row[0], row[1]
            if label_cell.value and str(label_cell.value).strip() == "Producto":
                result["producto"] = str(value_cell.value).strip() if value_cell.value else ""
                break

    if "Fórmula" in wb.sheetnames:
        ws_formula = wb["Fórmula"]
        wb_formulas = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
        ws_formula_raw = wb_formulas["Fórmula"]

        # --- Ubicar columnas de la tabla principal (Ingrediente/% /Función/INS) ---
        # por el texto de su encabezado, no por posición fija — distintos
        # Dossiers insertan columnas extra (ej. "Código SAP") que corren
        # todo lo demás.
        fila_header_principal, cols_principal = _buscar_encabezados(
            ws_formula,
            ["ingredientes / aditivos", "%", "funcion aditivos", "ins"],
        )
        col_ing = cols_principal.get("ingredientes / aditivos", 1)
        col_pct = cols_principal.get("%", 4)
        col_funcion = cols_principal.get("funcion aditivos", 5)
        col_ins = cols_principal.get("ins", 6)
        fila_inicio_principal = (fila_header_principal or 3) + 1

        # --- Ubicar la tabla de orden decreciente ("Fórmula con orden
        # decreciente de ingredientes") — puede estar en cualquier columna
        # a la derecha de la tabla principal según el Dossier.
        fila_titulo_decreciente, cols_titulo = _buscar_encabezados(
            ws_formula, ["formula con orden decreciente de ingredientes"]
        )
        col_titulo_decreciente = cols_titulo.get("formula con orden decreciente de ingredientes")

        col_ing_decreciente = None
        col_pct_decreciente = None
        fila_subheader_decreciente = None
        if fila_titulo_decreciente and col_titulo_decreciente:
            # La fila siguiente al título trae los sub-encabezados
            # "Ingrediente" / "%" — se buscan cerca de esa columna para no
            # confundirlas con la tabla principal (que también tiene "%").
            for fila_candidata in range(fila_titulo_decreciente + 1, fila_titulo_decreciente + 4):
                for cell in ws_formula[fila_candidata]:
                    if cell.column < col_titulo_decreciente - 2:
                        continue
                    texto = _normalize(cell.value)
                    if texto.startswith("ingrediente") and col_ing_decreciente is None:
                        col_ing_decreciente = cell.column
                        fila_subheader_decreciente = fila_candidata
                    elif texto == "%" and col_pct_decreciente is None:
                        col_pct_decreciente = cell.column
                if col_ing_decreciente and col_pct_decreciente:
                    break

        # --- Mapa nombre normalizado → % (desde la tabla principal) ---
        # Se usa como respaldo cuando el cruce por fórmula/caché no resuelve.
        # El valor queda normalizado a "porcentaje plano" (ver
        # _valor_como_porcentaje) para no mezclar convenciones distintas.
        pct_por_ingrediente = {}
        r = fila_inicio_principal
        while True:
            ing_a = ws_formula.cell(row=r, column=col_ing).value
            if ing_a is None:
                break
            pct_d = _valor_como_porcentaje(ws_formula.cell(row=r, column=col_pct))
            if pct_d is not None:
                pct_por_ingrediente[_normalize(ing_a)] = pct_d
            r += 1

        # --- Orden decreciente, con % resuelto en tres pasadas:
        # 1) si la celda es una fórmula que suma referencias a otras
        #    celdas, se usa el valor YA CALCULADO de esas celdas (sin
        #    asumir a qué columna de apoyo apuntan — varía según Dossier);
        # 2) si no es fórmula (o no se pudo resolver), se usa el valor
        #    cacheado de la propia celda;
        # 3) como último respaldo, se cruza por nombre contra la tabla
        #    principal.
        formula_rows = []
        ingredientes_orden = []
        if col_ing_decreciente and col_pct_decreciente:
            r = (fila_subheader_decreciente or 4) + 1
            while True:
                ing = ws_formula.cell(row=r, column=col_ing_decreciente).value
                if ing is None:
                    break
                ing_clean = str(ing).strip()
                if _normalize(ing_clean) == "total":
                    break
                celda_decreciente = ws_formula.cell(row=r, column=col_pct_decreciente)
                celda_formula = ws_formula_raw.cell(row=r, column=col_pct_decreciente).value
                pct_resuelto = _resolve_formula_refs(celda_formula, ws_formula)
                if pct_resuelto is not None:
                    # Resuelto sumando celdas referenciadas por la fórmula
                    # (valores crudos, sin normalizar) — se escala según el
                    # formato de ESTA celda (la que se muestra), no el de
                    # las celdas referenciadas.
                    if "%" in (celda_decreciente.number_format or ""):
                        pct_val = pct_resuelto * 100
                    else:
                        pct_val = pct_resuelto
                else:
                    # Sin fórmula resoluble: usar el valor cacheado de la
                    # propia celda (ya normalizado a porcentaje plano), o
                    # como último respaldo, cruzar por nombre.
                    pct_val = _valor_como_porcentaje(celda_decreciente)
                    if pct_val is None:
                        pct_val = pct_por_ingrediente.get(_normalize(ing_clean))
                pct_str = _fmt_pct(pct_val / 100) if isinstance(pct_val, (int, float)) else ""
                formula_rows.append(f"{ing_clean} | {pct_str}")
                ingredientes_orden.append(ing_clean)
                r += 1
        result["formula_rows"] = "\n".join(formula_rows)

        # --- Mapa nombre normalizado → función y → INS (tabla principal) ---
        funcion_por_ingrediente = {}
        ins_por_ingrediente = {}
        r = fila_inicio_principal
        while True:
            ing = ws_formula.cell(row=r, column=col_ing).value
            if ing is None:
                break
            funcion = ws_formula.cell(row=r, column=col_funcion).value
            ins = ws_formula.cell(row=r, column=col_ins).value
            if funcion and _normalize(funcion) not in ("n/a", "na", ""):
                funcion_por_ingrediente[_normalize(ing)] = str(funcion).strip()
            ins_fmt = _fmt_ins(ins)
            if ins_fmt:
                ins_por_ingrediente[_normalize(ing)] = ins_fmt
            r += 1
        # "Saborizantes naturales" ya viene agrupado en la tabla de orden
        # decreciente; si no está mapeado individualmente en la tabla
        # principal (porque ahí está desglosado en aromas), asignamos su
        # función fija conocida.
        funcion_por_ingrediente.setdefault(_normalize("Saborizantes naturales"), "Saborización")
        funcion_por_ingrediente.setdefault(_normalize("Sabores Naturales"), "Saborización")

        # --- Cruce: ingredientes (orden decreciente) + aditivos + filas ---
        aditivos_list = []
        aditivos_filas = []
        for idx, nombre in enumerate(ingredientes_orden, start=1):
            funcion = _buscar_en_mapa(nombre, funcion_por_ingrediente)
            if funcion:
                ins = _buscar_en_mapa(nombre, ins_por_ingrediente)
                nombre_fmt = f"{nombre.upper()} (INS {ins})" if ins else nombre.upper()
                aditivos_list.append(f"{nombre_fmt} | {funcion}")
                aditivos_filas.append(str(idx))

        result["ingredientes"] = "\n".join(ingredientes_orden)
        result["aditivos"] = "\n".join(aditivos_list)
        result["aditivos_filas"] = ", ".join(aditivos_filas)

    # --- Información nutricional (hoja de "Proyecto de Rótulo", columna 100 g/ml) ---
    # Esta hoja no siempre se llama "Proyecto de rótulo" en la pestaña —
    # algunos Dossiers la nombran con un código interno (ej. "PR PO8.08
    # CL"), y el texto "Proyecto Rótulo" aparece solo en el título dentro
    # del documento, no en el nombre de la pestaña. Por eso se busca
    # primero por nombre de pestaña (más rápido) y, si no aparece, por el
    # texto del título dentro de las primeras filas de cada hoja.
    def _es_hoja_proyecto_rotulo(ws):
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=15):
            for cell in row:
                if not cell.value:
                    continue
                texto = _normalize(cell.value)
                if "proyecto" in texto and "rotulo" in texto:
                    return True
        return False

    rotulo_sheet_name = next(
        (s for s in wb.sheetnames if _normalize(s).startswith("proyecto de rotulo")),
        None
    )
    if rotulo_sheet_name is None:
        rotulo_sheet_name = next(
            (s for s in wb.sheetnames if _es_hoja_proyecto_rotulo(wb[s])),
            None
        )
    if rotulo_sheet_name:
        ws_rotulo = wb[rotulo_sheet_name]
        # Mapa etiqueta normalizada → campo del template
        nut_map = {
            "energia (kcal)": "energia",
            "proteinas (g)": "proteina",
            "grasas totales (g)": "grasa_total",
            "grasa total (g)": "grasa_total",
            "grasas saturadas (g)": "grasa_sat",
            "grasas monoinsat. (g)": "grasa_mono",
            "grasas poliinsat. (g)": "grasa_poli",
            "grasas trans (g)": "grasa_trans",
            "colesterol (mg)": "colesterol",
            "carbohidratos totales (g)": "carb_totales",
            "carbohidratos disp. (g)": "carb_disp",
            "hidratos de carbono disponibles (g)": "carb_disp",
            "azucares totales (g)": "azucares",
            "fibra dietetica total (g)": "fibra",
            "sodio (mg)": "sodio",
        }
        # Buscar la columna "100 ml" o "100 g" — suele estar en una fila
        # encabezado cercana a "INFORMACIÓN NUTRICIONAL"
        col_100 = None
        for row in ws_rotulo.iter_rows(min_row=1, max_row=ws_rotulo.max_row):
            for cell in row:
                if cell.value and _normalize(cell.value) in ("100 ml", "100 g", "100g", "100ml"):
                    col_100 = cell.column
                    break
            if col_100:
                break
        if col_100:
            # La columna donde están las etiquetas ("Energía (kcal)", etc.)
            # no siempre es la B — varía según el Dossier. Se ubica
            # dinámicamente como la misma columna donde aparece el título
            # "INFORMACIÓN NUTRICIONAL" (las etiquetas de la tabla quedan
            # alineadas debajo de ese título en la práctica).
            col_label = None
            fila_inicio = None
            fila_fin = None
            for row in ws_rotulo.iter_rows(min_row=1, max_row=ws_rotulo.max_row, max_col=col_100):
                for cell in row:
                    if not cell.value:
                        continue
                    texto = _normalize(cell.value)
                    # Match exacto: el título de sección "2. INFORMACIÓN
                    # NUTRICIONAL" también contiene esta frase como
                    # substring, pero no es la celda que encabeza la
                    # tabla en sí (que repite la frase sola, sin el "2.").
                    if texto == "informacion nutricional" and fila_inicio is None:
                        fila_inicio = cell.row
                        col_label = cell.column
                if fila_inicio is not None:
                    # Ya encontramos el título; ahora buscamos el cierre
                    # "(*) En relación..." en esa misma columna de etiquetas.
                    label_cell = row[col_label - 1] if col_label and len(row) >= col_label else None
                    if label_cell and label_cell.value and _normalize(label_cell.value).startswith("(*) en relacion") and fila_fin is None:
                        fila_fin = label_cell.row
                        break
            if fila_inicio is None:
                fila_inicio = 1
            if col_label is None:
                col_label = 2
            if fila_fin is None:
                fila_fin = ws_rotulo.max_row

            micronutrientes_extra = []
            for row in ws_rotulo.iter_rows(min_row=fila_inicio, max_row=fila_fin, max_col=col_100):
                label_cell = row[col_label - 1] if len(row) >= col_label else None
                if label_cell is None or not label_cell.value:
                    continue
                label_norm = _normalize(label_cell.value)
                campo = nut_map.get(label_norm)
                val_cell = ws_rotulo.cell(row=label_cell.row, column=col_100)
                if campo:
                    if campo not in result and isinstance(val_cell.value, (int, float)):
                        result[campo] = str(val_cell.value)
                elif isinstance(val_cell.value, (int, float)):
                    nombre_limpio = str(label_cell.value).strip()
                    micronutrientes_extra.append(f"{nombre_limpio} | {val_cell.value}")
            if micronutrientes_extra:
                result["micronutrientes_extra"] = "\n".join(micronutrientes_extra)

    return result


# ── Test rápido ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else \
        "/mnt/user-data/uploads/Dossier_NotMilk_Original_10001202.xlsx"

    with open(path, "rb") as f:
        data = extract_dossier(f.read())

    print("Producto:", data.get("producto"))
    print("\nFormula rows:")
    print(data.get("formula_rows"))
    print("\nIngredientes (orden decreciente):")
    print(data.get("ingredientes"))
    print("\nAditivos:")
    print(data.get("aditivos"))
    print("\nAditivos filas:")
    print(data.get("aditivos_filas"))
